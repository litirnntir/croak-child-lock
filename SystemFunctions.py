import json
import os
import signal
import subprocess
import sys

import openpyxl as openpyxl
import osascript

from cryptography.fernet import Fernet


def format_time(seconds: int) -> str:
    """Форматирует время в секундах в виде ЧЧ:ММ:СС.

    Аргументы:
        seconds: время в секундах, целое число.

    Возвращает:
        Строку, содержащую время в формате ЧЧ:ММ:СС.
    """
    return f"{divmod(divmod(seconds, 3600)[0], 24)[1]:02}:{divmod(divmod(seconds, 3600)[1], 60)[0]:02}:{divmod(divmod(seconds, 3600)[1], 60)[1]:02}"


def resource_path(relative_path: str) -> str:
    user_dir = os.path.expanduser("~")
    app_dir = os.path.join(user_dir, "Croak")
    settings_file = os.path.join(app_dir, relative_path)
    return settings_file


with open(resource_path("key.key"), "rb") as file:
    key = file.read()

fernet = Fernet(key)


def encrypt_json(filename: str) -> None:
    """Шифрует содержимое JSON-файла с помощью ключа fernet.

    :param filename: имя JSON-файла для шифрования
    """
    with open(filename, "rb") as file:
        data = file.read()
    encrypted_data = fernet.encrypt(data)
    with open(filename, "wb") as file:
        file.write(encrypted_data)


def decrypt_json(filename: str) -> str:
    """Расшифровывает содержимое JSON-файла с помощью ключа fernet.

    :param filename: имя JSON-файла для расшифровки
    :return: расшифрованное содержимое JSON-файла в виде строки
    """
    with open(filename, "rb") as file:
        encrypted_data = file.read()
    decrypted_data = fernet.decrypt(encrypted_data)
    return decrypted_data.decode("utf-8")


def update_json(filename: str, key: str, value: any) -> None:
    """Обновляет значение по ключу в JSON-файле и шифрует его с помощью ключа fernet.

    :param filename: имя JSON-файла для обновления
    :param key: ключ, по которому нужно обновить значение
    :param value: новое значение для ключа
    """
    data = json.loads(decrypt_json(filename))
    data[key] = value
    # dict -> json -> bytes
    data = json.dumps(data)
    bytes_obj = data.encode('utf-8')
    encrypted_data = fernet.encrypt(bytes_obj)
    with open(filename, "wb") as f:
        f.write(encrypted_data)


def update_json_without_encrypt(filename: str, text_json: str) -> None:
    """Обновляет значение в JSON-файле

    :param filename: имя JSON-файла для обновления
    :param text_json: строка, записываемая в файл
    """
    with open(filename, "wb") as f:
        f.write(text_json)


def delete_from_json(filename: str, key: str) -> None:
    """Удаляет значение по ключу в JSON-файле и шифрует его с помощью ключа fernet.

    :param filename: имя JSON-файла для удаления
    :param key: ключ, по которому нужно удалить значение
    """
    data = json.loads(decrypt_json(filename))
    if key in data:
        del data[key]
        # dict -> json -> bytes
        data = json.dumps(data)
        bytes_obj = data.encode('utf-8')
        encrypted_data = fernet.encrypt(bytes_obj)
        with open(filename, "wb") as f:
            f.write(encrypted_data)


def get_from_json(filename: str) -> dict:
    """Дешифрует json файл и возвращает данные

    :param filename: имя JSON-файла для чтения
    """
    return json.loads(decrypt_json(filename))


def get_from_json_without_encrypt(filename: str) -> str:
    """Открывает json файл и возвращает данные

    :param filename: имя JSON-файла для чтения
    """
    with open(filename, "r") as f:
        return f.read()


def reset_json(filename: str) -> None:
    """Очищает json файл и снова шифрует его"""
    with open(filename, "w") as f:
        f.write("{}")
    encrypt_json(filename)


def get_active_app_name() -> str:
    """Возвращает имя активного приложения на Mac OS"""
    script = """
    tell application "System Events"
        set frontApp to first application process whose frontmost is true
        set appPath to POSIX path of (frontApp's application file)
    end tell
    return appPath
    """
    output = subprocess.check_output(["osascript", "-e", script])
    folder_name = output.strip().decode("utf-8").split("/")[-1].replace(".app", "")
    return folder_name


def send_notification(text: str) -> None:
    """Отправляет уведомление на Mac OS с заданным текстом и заголовком "Croak" """
    osascript.run("defaults write com.apple.notificationcenterui bannerTime 2")
    command = f'display notification "{text}" with title "Croak"'
    osascript.run(command)


def apps_list() -> list[str]:
    """Возвращает список всех приложений в папке /Applications"""
    apps = []
    app_path = "/Applications"
    for file in os.listdir(app_path):
        if file.endswith(".app"):
            apps.append(file[:-4])
    return sorted(apps)


def close_app(app_name: str) -> None:
    """Закрывает приложение по его имени, если оно запущено"""
    try:
        processes = os.popen("ps ax").readlines()
        for process in processes:
            if app_name in process:
                fields = process.split()
                pid = fields[0]
                os.kill(int(pid), signal.SIGTERM)
                print(f"Закрыли приложение {app_name}")
    except:
        pass


def get_open_apps() -> list[str]:
    """Возвращает список всех открытых приложений на Mac OS"""
    script = 'tell application "System Events" to get name of every process whose background only is false'
    output = subprocess.check_output(['osascript', '-e', script])
    output = output.decode('utf-8').strip().split(', ')
    return output


def save_stats_to_file(directory: str, stats_data: dict[str, any]) -> None:
    """Сохраняет статистику по приложениям в файл Excel.

    Аргументы:
        directory: путь к файлу, в который нужно сохранить статистику.
        stats_data: словарь, в котором ключи - названия приложений, а значения - время использования в секундах.

    Возвращает:
        None
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    data = stats_data

    row = 1
    for app, time in data.items():
        ws.cell(row=row, column=1).value = app
        ws.cell(row=row,
                column=2).value = format_time(time)
        row += 1

    wb.save(directory)
